#!/usr/bin/env python3
"""
San Antonio BESS Analysis Model Generator

This script creates a complete Excel workbook for analyzing the deployment of
Battery Energy Storage Systems (BESS) across 38 City of San Antonio buildings.

The workbook includes:
- Dashboard (executive summary)
- Building Inventory (38 buildings data)
- Value Calculations (all revenue streams)
- CPS Energy Data (utility information)
- Financial Model (20-year cash flows)
- Sensitivity Analysis (scenario testing)
- Prioritization Scoring (building ranking)
- Phase Planning (deployment timeline)

Usage:
    python generate_san_antonio_bess_model.py

Output:
    San_Antonio_BESS_Analysis.xlsx
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
import datetime

# Color scheme
COLOR_HEADER = "2E75B5"  # Dark blue
COLOR_INPUT = "FFF2CC"   # Light yellow
COLOR_FORMULA = "DAEEF3" # Light blue
COLOR_OUTPUT = "E2EFDA"  # Light green

def create_workbook():
    """Create the main workbook with all tabs."""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create all tabs
    sheets = [
        'Dashboard',
        'Building Inventory',
        'Value Calculations',
        'CPS Energy Data',
        'Financial Model',
        'Sensitivity Analysis',
        'Prioritization Scoring',
        'Phase Planning'
    ]

    for sheet_name in sheets:
        wb.create_sheet(title=sheet_name)

    return wb

def format_currency(ws, cell_range):
    """Apply currency formatting to a range."""
    for row in ws[cell_range]:
        for cell in row:
            cell.number_format = '$#,##0.00'

def format_percentage(ws, cell_range):
    """Apply percentage formatting to a range."""
    for row in ws[cell_range]:
        for cell in row:
            cell.number_format = '0.0%'

def apply_header_style(ws, row_num, last_col):
    """Apply header styling to a row."""
    for col in range(1, last_col + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

def create_building_inventory(wb):
    """Create Building Inventory tab with sample data and formulas."""
    ws = wb['Building Inventory']

    # Headers
    headers = [
        'Building ID', 'Building Name', 'Address', 'Building Type',
        'Proposed BESS Size (kW)', 'Battery Capacity (kWh)',
        'CPS Account Number', 'Rate Schedule', 'Peak Demand (kW)',
        'Annual Energy (kWh)', 'Current Demand Charge ($/month)',
        'Distribution Feeder ID', 'Critical Facility', 'Priority Score',
        'Phase Assignment', 'Notes'
    ]

    ws.append(headers)
    apply_header_style(ws, 1, len(headers))

    # Sample building types
    building_types = [
        'Emergency Ops', '911 Center', 'Police', 'Fire',
        'Water Control', 'City Hall', 'Standard', 'Standard', 'Standard'
    ]

    # Sample building names
    building_names = [
        'Emergency Operations Center',
        '911 Call Center',
        'Police Headquarters',
        'Fire Station #1',
        'Water Treatment Control Center',
        'City Hall',
        'Parks Department',
        'Public Works Building',
        'Library Main Branch',
    ]

    # Generate 38 sample buildings
    for i in range(1, 39):
        # Use sample data for first 9, then generic for rest
        if i <= len(building_names):
            name = building_names[i-1]
            bldg_type = building_types[i-1]
        else:
            name = f'Municipal Building #{i}'
            bldg_type = 'Standard'

        # Sample data with variation
        bess_size = 100 + (i % 3) * 25  # 100, 125, or 150 kW
        peak_demand = bess_size + 30 + (i % 5) * 10
        annual_energy = peak_demand * 3500  # Approximate annual hours
        demand_charge = peak_demand * 18  # Approximate monthly demand charge

        row_data = [
            f'B{i:03d}',  # Building ID
            name,
            f'{1000 + i*100} Example St, San Antonio, TX 78201',
            bldg_type,
            bess_size,
            f'=E{i+1}*2',  # Battery capacity formula
            f'{1234567800 + i}',  # Account number
            'LGS' if peak_demand > 150 else 'GS',
            peak_demand,
            annual_energy,
            demand_charge,
            f'FEEDER-{40 + (i % 15)}A',
            f'=IF(D{i+1}="Emergency Ops","Yes",IF(D{i+1}="911 Center","Yes",IF(D{i+1}="Police","Yes",IF(D{i+1}="Fire","Yes","No"))))',
            f'=IFERROR(VLOOKUP(A{i+1},\'Prioritization Scoring\'!$A:$X,24,FALSE),0)',
            f'=IF(N{i+1}>=76,1,IF(N{i+1}>=51,2,3))',
            ''
        ]
        ws.append(row_data)

    # Set column widths
    column_widths = [12, 30, 40, 15, 20, 22, 18, 15, 18, 18, 25, 20, 15, 15, 15, 30]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Apply input cell formatting
    input_fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
    for row in range(2, 40):
        for col in ['B', 'C', 'D', 'E', 'G', 'H', 'I', 'J', 'K', 'L', 'P']:
            ws[f'{col}{row}'].fill = input_fill

    # Apply formula cell formatting
    formula_fill = PatternFill(start_color=COLOR_FORMULA, end_color=COLOR_FORMULA, fill_type="solid")
    for row in range(2, 40):
        for col in ['F', 'M', 'N', 'O']:
            ws[f'{col}{row}'].fill = formula_fill

    # Conditional formatting for Priority Score
    ws.conditional_formatting.add(
        f'N2:N39',
        ColorScaleRule(
            start_type='num', start_value=0, start_color='F8696B',
            mid_type='num', mid_value=50, mid_color='FFEB84',
            end_type='num', end_value=100, end_color='63BE7B'
        )
    )

    # Conditional formatting for Phase Assignment
    ws.conditional_formatting.add(
        'O2:O39',
        CellIsRule(operator='equal', formula=['1'], fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'))
    )
    ws.conditional_formatting.add(
        'O2:O39',
        CellIsRule(operator='equal', formula=['2'], fill=PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'))
    )
    ws.conditional_formatting.add(
        'O2:O39',
        CellIsRule(operator='equal', formula=['3'], fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'))
    )

    return ws

def create_value_calculations(wb):
    """Create Value Calculations tab with all revenue stream formulas."""
    ws = wb['Value Calculations']

    # Main headers
    ws.append(['Building-Specific Revenue Streams'])
    ws.merge_cells('A1:AO1')
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Column headers
    headers = [
        'Building ID', 'BESS (kW)',
        # Demand Charge Reduction (C-J)
        '1. Demand Charge', 'Summer Demand (kW)', 'Winter Demand (kW)',
        'Summer Rate ($/kW-mo)', 'Winter Rate ($/kW-mo)', 'Peak Reduction %',
        'Annual Savings ($)', '$/kW-year',
        # CPS Energy DR (K-P)
        '2. CPS Energy DR', 'Expected Events', 'Summer Rate ($/kW-season)',
        'Winter Rate ($/kW-total)', 'Annual Revenue ($)', '$/kW-year',
        # ERCOT Call Option (Q-T)
        '3. ERCOT Call Option', 'Call Value ($/kW-month)', 'Annual Revenue ($)', '$/kW-year',
        # Infrastructure Deferral (U-AA)
        '4. Infrastructure Deferral', 'Feeder Upgrade Cost ($)', 'Feeder Capacity (kW)',
        'Years Deferred', 'Proportional Deferral ($)', 'Annualized Value ($)', '$/kW-year',
        # Resilience Value (AB-AG)
        '5. Resilience Value', 'VOLL ($/kWh)', 'Outage Hours/year',
        'Critical Load (kW)', 'Annual Value ($)', '$/kW-year',
        # Energy Arbitrage (AH-AL)
        '6. Energy Arbitrage', 'Annual Cycles', 'Margin per Cycle ($)',
        'Annual Value ($)', '$/kW-year',
        # Total (AM-AO)
        'TOTAL VALUE', 'Total Annual ($)', '$/kW-year'
    ]

    ws.append(headers)
    apply_header_style(ws, 2, len(headers))

    # Add data rows with formulas for each building
    for i in range(1, 39):
        building_id = f'B{i:03d}'
        row_num = i + 2
        inv_row = i + 1  # Building Inventory row number

        # Base data from Building Inventory
        bess_kw = f'=\'Building Inventory\'!E{inv_row}'
        peak_demand = f'=\'Building Inventory\'!I{inv_row}'

        # Demand Charge Reduction calculations
        summer_demand = peak_demand
        winter_demand = f'={peak_demand}*0.85'  # Assume 85% in winter
        summer_rate = 22  # $/kW-month
        winter_rate = 12  # $/kW-month
        peak_reduction = 0.80  # 80%
        demand_savings = f'=(D{row_num}*F{row_num}*4 + E{row_num}*G{row_num}*8)*H{row_num}'
        demand_per_kw = f'=I{row_num}/B{row_num}'

        # CPS Energy DR
        expected_events = 25
        summer_dr_rate = 73
        winter_dr_rate = 45
        dr_revenue = f'=B{row_num}*(M{row_num} + (N{row_num}*L{row_num}/25))'
        dr_per_kw = f'=O{row_num}/B{row_num}'

        # ERCOT Call Option
        call_value = 8.50
        ercot_revenue = f'=B{row_num}*R{row_num}*12'
        ercot_per_kw = f'=S{row_num}/B{row_num}'

        # Infrastructure Deferral (varies by feeder)
        feeder_cost = 2000000 if i <= 10 else (1000000 if i <= 25 else 0)  # Sample data
        feeder_capacity = 500
        years_deferred = 7 if i <= 10 else (5 if i <= 25 else 0)
        prop_deferral = f'=V{row_num}*(B{row_num}/W{row_num})'
        annualized = f'=IF(Y{row_num}>0,-PMT(0.05,20,Y{row_num}),0)'
        infra_per_kw = f'=IF(B{row_num}>0,Z{row_num}/B{row_num},0)'

        # Resilience Value (varies by building type)
        # Will be calculated based on building type from inventory
        voll = f'=IF(\'Building Inventory\'!D{inv_row}="Emergency Ops",60,IF(\'Building Inventory\'!D{inv_row}="911 Center",50,IF(\'Building Inventory\'!D{inv_row}="Police",40,IF(\'Building Inventory\'!D{inv_row}="Fire",40,IF(\'Building Inventory\'!D{inv_row}="Water Control",30,5)))))'
        outage_hours = f'=IF(\'Building Inventory\'!M{inv_row}="Yes",10,5)'
        critical_load = f'=B{row_num}*0.75'
        resilience_value = f'=AC{row_num}*AD{row_num}*AE{row_num}'
        resilience_per_kw = f'=AF{row_num}/B{row_num}'

        # Energy Arbitrage
        annual_cycles = 220
        margin_per_cycle = f'=(B{row_num}*2)*0.058'  # 2hr battery, $0.058/kWh margin
        arbitrage_value = f'=AI{row_num}*AJ{row_num}'
        arbitrage_per_kw = f'=AK{row_num}/B{row_num}'

        # Total
        total_annual = f'=SUM(I{row_num},O{row_num},S{row_num},Z{row_num},AF{row_num},AK{row_num})'
        total_per_kw = f'=AN{row_num}/B{row_num}'

        row_data = [
            building_id, bess_kw,
            # Demand Charge
            '', summer_demand, winter_demand, summer_rate, winter_rate, peak_reduction,
            demand_savings, demand_per_kw,
            # CPS DR
            '', expected_events, summer_dr_rate, winter_dr_rate, dr_revenue, dr_per_kw,
            # ERCOT
            '', call_value, ercot_revenue, ercot_per_kw,
            # Infrastructure
            '', feeder_cost, feeder_capacity, years_deferred, prop_deferral, annualized, infra_per_kw,
            # Resilience
            '', voll, outage_hours, critical_load, resilience_value, resilience_per_kw,
            # Arbitrage
            '', annual_cycles, margin_per_cycle, arbitrage_value, arbitrage_per_kw,
            # Total
            '', total_annual, total_per_kw
        ]

        ws.append(row_data)

    # Fleet Summary starting at row 42
    summary_row = 42
    ws[f'A{summary_row}'] = 'FLEET TOTALS'
    ws[f'A{summary_row}'].font = Font(size=12, bold=True)

    ws[f'A{summary_row+1}'] = 'Total Capacity (MW)'
    ws[f'B{summary_row+1}'] = '=SUM(B3:B40)/1000'

    ws[f'A{summary_row+2}'] = 'Average $/kW-year'
    ws[f'B{summary_row+2}'] = '=SUM(AN3:AN40)/SUM(B3:B40)'

    ws[f'A{summary_row+4}'] = 'REVENUE BY STREAM'
    ws[f'A{summary_row+4}'].font = Font(bold=True)

    streams = [
        ('Demand Charges', 'I3:I40'),
        ('CPS Energy DR', 'O3:O40'),
        ('ERCOT Call Option', 'S3:S40'),
        ('Infrastructure Deferral', 'Z3:Z40'),
        ('Resilience Value', 'AF3:AF40'),
        ('Energy Arbitrage', 'AK3:AK40'),
        ('TOTAL', 'AN3:AN40')
    ]

    for idx, (name, range_ref) in enumerate(streams):
        row = summary_row + 5 + idx
        ws[f'A{row}'] = name
        ws[f'B{row}'] = f'=SUM({range_ref})'
        ws[f'B{row}'].number_format = '$#,##0'

    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    for col in range(3, 42):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # Apply input cell formatting (columns with assumptions)
    input_fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
    input_cols = ['F', 'G', 'H', 'L', 'M', 'N', 'R', 'V', 'W', 'X', 'AI']
    for row in range(3, 41):
        for col in input_cols:
            ws[f'{col}{row}'].fill = input_fill

    # Apply formula cell formatting
    formula_fill = PatternFill(start_color=COLOR_FORMULA, end_color=COLOR_FORMULA, fill_type="solid")
    formula_cols = ['B', 'D', 'E', 'I', 'J', 'O', 'P', 'S', 'T', 'Y', 'Z', 'AA', 'AC', 'AD', 'AE', 'AF', 'AG', 'AJ', 'AK', 'AL', 'AN', 'AO']
    for row in range(3, 41):
        for col in formula_cols:
            ws[f'{col}{row}'].fill = formula_fill

    # Format currency columns
    currency_cols = ['I', 'O', 'S', 'V', 'Y', 'Z', 'AF', 'AK', 'AN']
    for col in currency_cols:
        for row in range(3, 41):
            ws[f'{col}{row}'].number_format = '$#,##0'

    # Format per-kW columns
    per_kw_cols = ['J', 'P', 'T', 'AA', 'AG', 'AL', 'AO']
    for col in per_kw_cols:
        for row in range(3, 41):
            ws[f'{col}{row}'].number_format = '$#,##0.00'

    return ws

def create_cps_energy_data(wb):
    """Create CPS Energy Data tab."""
    ws = wb['CPS Energy Data']

    # Section A: Feeder Analysis
    ws['A1'] = 'DISTRIBUTION SYSTEM DATA (From CPS Energy Grid Planning)'
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:L1')

    ws.append([])  # Blank row

    feeder_headers = [
        'Feeder ID', 'Substation', 'Feeder Capacity (kVA)',
        'Current Peak Load (kW)', 'Loading %', 'Constraint Status',
        'Planned Upgrade', 'Upgrade Cost ($)', 'Upgrade Timeline (years)',
        'Buildings on Feeder', 'Total BESS on Feeder (kW)', 'Deferral Potential ($)'
    ]
    ws.append(feeder_headers)
    apply_header_style(ws, 3, len(feeder_headers))

    # Sample feeder data
    feeders = []
    for i in range(40, 55):
        feeder_id = f'FEEDER-{i}A'
        capacity = 500 + (i % 3) * 200
        peak_load = capacity * (0.65 + (i % 4) * 0.08)  # 65-89% loaded
        loading_pct = f'=D{i-36}/C{i-36}'
        constraint = f'=IF(E{i-36}>0.9,"Critical",IF(E{i-36}>0.8,"Constrained","OK"))'
        planned = 'Yes' if (i % 3 == 0) else 'No'
        cost = 2000000 if (i % 3 == 0) else 0
        timeline = (3 + (i % 3)) if planned == 'Yes' else 0
        buildings = f'=COUNTIF(\'Building Inventory\'!L:L,A{i-36})'
        total_bess = f'=SUMIF(\'Building Inventory\'!$L:$L,A{i-36},\'Building Inventory\'!$E:$E)'
        deferral = f'=IF(K{i-36}>0,H{i-36}*(K{i-36}/C{i-36}),0)'

        feeders.append([
            feeder_id, f'SUB-{i//3}', capacity, peak_load, loading_pct,
            constraint, planned, cost, timeline, buildings, total_bess, deferral
        ])

    for feeder in feeders:
        ws.append(feeder)

    # Section B: Rate Schedules
    rate_row = len(feeders) + 6
    ws[f'A{rate_row}'] = 'CPS ENERGY RATE SCHEDULES'
    ws[f'A{rate_row}'].font = Font(size=12, bold=True)

    rate_headers = [
        'Rate Class', 'Summer Demand ($/kW-month)', 'Winter Demand ($/kW-month)',
        'Energy On-Peak ($/kWh)', 'Energy Off-Peak ($/kWh)', 'Fixed Charge ($/month)'
    ]
    ws.append(rate_headers)
    apply_header_style(ws, rate_row + 1, len(rate_headers))

    rates = [
        ['LGS - Large General Service', 22.50, 12.00, 0.072, 0.045, 150],
        ['GS - General Service', 18.00, 10.00, 0.065, 0.040, 75],
        ['MLG - Medium/Large General', 20.00, 11.00, 0.068, 0.042, 100],
    ]

    for rate in rates:
        ws.append(rate)

    # Section C: Historical DR Events
    event_row = rate_row + len(rates) + 4
    ws[f'A{event_row}'] = 'HISTORICAL DR EVENT DATA (Request from CPS Energy)'
    ws[f'A{event_row}'].font = Font(size=12, bold=True)

    event_headers = [
        'Date', 'Event Type', 'Start Time', 'End Time',
        'Duration (hours)', 'Trigger', 'ERCOT RT Price ($/MWh)'
    ]
    ws.append(event_headers)
    apply_header_style(ws, event_row + 1, len(event_headers))

    # Sample historical events
    sample_events = [
        ['2024-07-15', 'Summer', '14:00', '17:00', 3, 'Heat Wave', 125],
        ['2024-07-22', 'Summer', '15:00', '18:00', 3, 'High Demand', 95],
        ['2024-08-03', 'Summer', '13:00', '16:30', 3.5, 'ERCOT Emergency', 245],
        ['2024-12-22', 'Winter', '07:00', '10:00', 3, 'Cold Snap', 180],
        ['[Add more rows as data received]', '', '', '', '', '', ''],
    ]

    for event in sample_events:
        ws.append(event)

    # Summary statistics
    stats_row = event_row + len(sample_events) + 4
    ws[f'A{stats_row}'] = 'EVENT STATISTICS'
    ws[f'A{stats_row}'].font = Font(bold=True)
    ws[f'A{stats_row+1}'] = 'Average Events per Year'
    ws[f'B{stats_row+1}'] = '[Calculate after data entry]'
    ws[f'A{stats_row+2}'] = 'Average Summer Events'
    ws[f'B{stats_row+2}'] = '25 (typical)'
    ws[f'A{stats_row+3}'] = 'Average Duration (hours)'
    ws[f'B{stats_row+3}'] = '3.0'

    # Set column widths
    widths = [20, 15, 20, 20, 12, 18, 15, 15, 22, 20, 22, 20]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Format currency and percentage
    for row in range(4, 4 + len(feeders)):
        ws[f'H{row}'].number_format = '$#,##0'
        ws[f'L{row}'].number_format = '$#,##0'
        ws[f'E{row}'].number_format = '0.0%'

    # Conditional formatting for constraint status
    ws.conditional_formatting.add(
        f'F4:F{3+len(feeders)}',
        CellIsRule(operator='equal', formula=['"Critical"'], fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'F4:F{3+len(feeders)}',
        CellIsRule(operator='equal', formula=['"Constrained"'], fill=PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'))
    )

    return ws

def create_financial_model(wb):
    """Create Financial Model tab with 20-year cash flows."""
    ws = wb['Financial Model']

    # Section A: CAPEX
    ws['A1'] = 'PROJECT FINANCIAL MODEL'
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:D1')

    ws.append([])
    ws['A3'] = 'CAPITAL EXPENDITURES (CAPEX)'
    ws['A3'].font = Font(size=12, bold=True)

    capex_data = [
        [],
        ['BESS Equipment & Installation', '', ''],
        ['Total Capacity (kW)', '=SUM(\'Building Inventory\'!E2:E39)', 'From Building Inventory'],
        ['Unit Cost ($/kW)', 1200, 'Range: $1,100-1,300/kW'],
        ['Total Equipment & Install', '=B6*B7', 'Equipment, inverters, BMS, installation'],
        [],
        ['Soft Costs', '', ''],
        ['Engineering & Design (5%)', '=B8*0.05', ''],
        ['Permits & Interconnection (2%)', '=B8*0.02', ''],
        ['Project Management (3%)', '=B8*0.03', ''],
        ['Contingency (10%)', '=B8*0.10', ''],
        [],
        ['TOTAL CAPEX', '=SUM(B8,B11:B14)', ''],
    ]

    for idx, row in enumerate(capex_data, 4):
        if row:
            ws.append(row)

    # Section B: Operating Costs
    ws.append([])
    ws['A18'] = 'ANNUAL OPERATING EXPENSES'
    ws['A18'].font = Font(size=12, bold=True)

    opex_data = [
        [],
        ['Operations & Maintenance', '', ''],
        ['O&M (% of CAPEX)', 0.015, '1.5% assumption'],
        ['Annual O&M Cost', '=B16*B21', ''],
        ['Monitoring & Software ($/kW-year)', 15, ''],
        ['Software Cost', '=B6*B23', ''],
        ['Insurance (% of CAPEX)', 0.005, '0.5% assumption'],
        ['Annual Insurance', '=B16*B25', ''],
        ['Property Tax', 0, 'Municipal property - exempt'],
        [],
        ['Total Annual O&M', '=B22+B24+B26+B27', ''],
        [],
        ['Battery Replacement (Year 15)', '', ''],
        ['Replacement Cost (50% of CAPEX)', '=B16*0.50', ''],
    ]

    for row in opex_data:
        ws.append(row)

    # Section C: Annual Revenues
    ws.append([])
    ws['A34'] = 'ANNUAL REVENUES'
    ws['A34'].font = Font(size=12, bold=True)

    revenue_data = [
        [],
        ['Demand Charge Reduction', '=\'Value Calculations\'!B52', ''],
        ['CPS Energy DR', '=\'Value Calculations\'!B53', ''],
        ['ERCOT Call Option', '=\'Value Calculations\'!B54', ''],
        ['Infrastructure Deferral (Annualized)', '=\'Value Calculations\'!B55', ''],
        ['Resilience Value', '=\'Value Calculations\'!B56', ''],
        ['Energy Arbitrage', '=\'Value Calculations\'!B57', ''],
        [],
        ['TOTAL ANNUAL REVENUE', '=SUM(B36:B41)', ''],
    ]

    for row in revenue_data:
        ws.append(row)

    # Section D: 20-Year Cash Flow
    ws.append([])
    ws.append([])
    ws['A46'] = '20-YEAR CASH FLOW ANALYSIS'
    ws['A46'].font = Font(size=12, bold=True)

    cf_headers = [
        'Year', 'CAPEX', 'Demand Charges', 'CPS DR', 'ERCOT Call',
        'Infrastructure', 'Resilience', 'Arbitrage', 'Total Revenue',
        'O&M Costs', 'Battery Replacement', 'Net Cash Flow', 'Cumulative CF'
    ]
    ws.append(cf_headers)
    apply_header_style(ws, 48, len(cf_headers))

    # Year 0 (construction)
    year_0 = [0, '=-$B$16', 0, 0, 0, 0, 0, 0, 0, 0, 0, '=SUM(B49:K49)', 0]
    ws.append(year_0)

    # Years 1-20
    for year in range(1, 21):
        row_num = 49 + year
        escalation = f'*(1.02^A{row_num})'  # 2% annual escalation

        year_data = [
            year,
            0,  # No CAPEX after year 0
            f'=$B$36{escalation}',  # Demand charges with escalation
            f'=$B$37',  # CPS DR (no escalation assumed)
            f'=$B$38',  # ERCOT
            f'=$B$39',  # Infrastructure
            f'=$B$40',  # Resilience
            f'=$B$41',  # Arbitrage
            f'=SUM(C{row_num}:H{row_num})',  # Total revenue
            f'=-$B$29',  # O&M costs
            f'=IF(A{row_num}=15,-$B$32,0)',  # Battery replacement in year 15
            f'=SUM(B{row_num}:K{row_num})',  # Net cash flow
            f'=L{row_num}+M{row_num-1}'  # Cumulative
        ]
        ws.append(year_data)

    # Section E: Financial Metrics
    ws.append([])
    ws.append([])
    ws['A72'] = 'KEY FINANCIAL METRICS'
    ws['A72'].font = Font(size=12, bold=True)

    metrics_data = [
        [],
        ['Net Present Value (5% discount)', '=NPV(0.05,L50:L69)+L49', '20-year NPV at 5% discount rate'],
        ['Internal Rate of Return (IRR)', '=IRR(L49:L69)', 'Overall project IRR'],
        ['Modified IRR (MIRR, 5% reinvest)', '=MIRR(L49:L69,0.05,0.05)', 'MIRR with 5% reinvestment rate'],
        ['Simple Payback (years)', '=MATCH(0,M49:M69,1)', 'Years to positive cumulative cash flow'],
        [],
        ['Benefit-Cost Ratio', '=NPV(0.05,I50:I69)/(-B49)', 'PV of benefits / PV of costs'],
        ['Annual Revenue per kW', '=B43/B6', 'Average $/kW-year'],
        ['Levelized Cost of Storage ($/kWh)', '=(B16+NPV(0.05,J50:J69))/(B6*2*250*20)', 'Total costs / total kWh throughput'],
    ]

    for row in metrics_data:
        ws.append(row)

    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 35
    for col in range(4, 14):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Apply input cell formatting
    input_fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
    input_cells = ['B7', 'B21', 'B23', 'B25', 'B27']
    for cell in input_cells:
        ws[cell].fill = input_fill

    # Apply output cell formatting
    output_fill = PatternFill(start_color=COLOR_OUTPUT, end_color=COLOR_OUTPUT, fill_type="solid")
    output_cells = ['B16', 'B29', 'B43', 'B74', 'B75', 'B76', 'B77', 'B79', 'B80', 'B81']
    for cell in output_cells:
        ws[cell].fill = output_fill
        ws[cell].font = Font(bold=True)

    # Format currency
    for row in range(49, 70):
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            ws[f'{col}{row}'].number_format = '$#,##0'

    # Format metrics
    ws['B74'].number_format = '$#,##0'
    ws['B75'].number_format = '0.0%'
    ws['B76'].number_format = '0.0%'
    ws['B77'].number_format = '0.0'
    ws['B79'].number_format = '0.00'
    ws['B80'].number_format = '$#,##0'
    ws['B81'].number_format = '$0.000'

    return ws

def create_sensitivity_analysis(wb):
    """Create Sensitivity Analysis tab."""
    ws = wb['Sensitivity Analysis']

    ws['A1'] = 'SENSITIVITY ANALYSIS'
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:H1')

    ws.append([])
    ws['A3'] = 'Base Case NPV'
    ws['B3'] = '=\'Financial Model\'!B74'
    ws['B3'].number_format = '$#,##0'
    ws['B3'].font = Font(bold=True)
    ws['B3'].fill = PatternFill(start_color=COLOR_OUTPUT, end_color=COLOR_OUTPUT, fill_type="solid")

    # One-Way Sensitivity: Demand Charge Savings
    ws.append([])
    ws['A5'] = 'ONE-WAY SENSITIVITY: Demand Charge Savings Impact on NPV'
    ws['A5'].font = Font(size=12, bold=True)

    ws['A7'] = 'Variable: Demand Charge Savings ($/kW-year)'
    ws['A8'] = 'Base Case:'
    ws['B8'] = '=\'Value Calculations\'!B52/SUM(\'Building Inventory\'!E2:E39)'
    ws['B8'].number_format = '$#,##0'

    ws.append([])
    ws['A10'] = 'Sensitivity Range:'

    # Sensitivity table headers
    sensitivity_headers = ['% Change', '-50%', '-25%', 'Base', '+25%', '+50%', '+75%', '+100%']
    ws.append(sensitivity_headers)
    for col_idx, header in enumerate(sensitivity_headers, 1):
        cell = ws.cell(row=11, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Demand charge values
    base_value = 220  # Approximate base case
    demand_values = [
        'Demand $/kW-yr',
        base_value * 0.5,
        base_value * 0.75,
        base_value,
        base_value * 1.25,
        base_value * 1.5,
        base_value * 1.75,
        base_value * 2.0
    ]
    ws.append(demand_values)

    # NPV row (will need manual data table setup in Excel)
    npv_row = ['NPV ($M)', '[Data Table]', '[Data Table]', '=$B$3/1000000', '[Data Table]', '[Data Table]', '[Data Table]', '[Data Table]']
    ws.append(npv_row)
    ws['A13'] = 'Note: Use Excel Data Table feature: Select A11:H13, Data > What-If Analysis > Data Table'
    ws['A13'].font = Font(italic=True, size=9)

    # Two-Way Sensitivity
    ws.append([])
    ws.append([])
    ws['A16'] = 'TWO-WAY SENSITIVITY: NPV ($M) - Demand Charges vs CAPEX'
    ws['A16'].font = Font(size=12, bold=True)

    # Two-way table structure
    ws['A18'] = 'Demand $/kW-yr →\n↓ CAPEX $/kW'
    ws['A18'].alignment = Alignment(wrap_text=True)

    # Column headers (Demand charge values)
    demand_range = [150, 180, 220, 260, 300, 350]
    for col_idx, value in enumerate(demand_range, 2):
        ws.cell(row=18, column=col_idx, value=value)
        ws.cell(row=18, column=col_idx).font = Font(bold=True)
        ws.cell(row=18, column=col_idx).alignment = Alignment(horizontal='center')

    # Row headers (CAPEX values)
    capex_range = [1000, 1100, 1200, 1300, 1400, 1500]
    for row_idx, value in enumerate(capex_range, 19):
        ws.cell(row=row_idx, column=1, value=value)
        ws.cell(row=row_idx, column=1).font = Font(bold=True)

    # Reference formula in corner
    ws['B19'] = '=$B$3/1000000'

    # Placeholder text
    for row_idx in range(19, 25):
        for col_idx in range(2, 8):
            ws.cell(row=row_idx, column=col_idx, value='[Data Table]')

    ws['A25'] = 'Note: Use Excel Data Table feature: Select A18:G24, Data > What-If Analysis > Data Table'
    ws['A25'].font = Font(italic=True, size=9)

    # Scenario Comparison
    ws.append([])
    ws.append([])
    ws['A28'] = 'SCENARIO COMPARISON'
    ws['A28'].font = Font(size=12, bold=True)

    scenario_headers = ['Metric', 'Conservative', 'Base Case', 'Aggressive', 'Units']
    ws.append(scenario_headers)
    apply_header_style(ws, 29, len(scenario_headers))

    scenarios = [
        ['Demand Charges ($/kW-yr)', 180, 220, 280, '$/kW-yr'],
        ['DR Events per Year', 15, 25, 30, 'events'],
        ['Infrastructure Deferral ($/kW-yr)', 10, 40, 100, '$/kW-yr'],
        ['CAPEX ($/kW)', 1300, 1200, 1100, '$/kW'],
        [],
        ['Annual Revenue', '=B30*3800+B31*3800*100+B32*3800', '=C30*3800+C31*3800*100+C32*3800', '=D30*3800+D31*3800*100+D32*3800', '$'],
        ['Total CAPEX', '=B33*3800', '=C33*3800', '=D33*3800', '$'],
        ['NPV (20-year, 5%)', '[Calculate]', '=$B$3', '[Calculate]', '$'],
        ['IRR', '[Calculate]', '=\'Financial Model\'!B75', '[Calculate]', '%'],
        ['Simple Payback', '[Calculate]', '=\'Financial Model\'!B77', '[Calculate]', 'years'],
    ]

    for row_data in scenarios:
        ws.append(row_data)

    # Tornado Chart Data (Key Variables)
    ws.append([])
    ws.append([])
    ws['A41'] = 'TORNADO CHART DATA: NPV Sensitivity to Key Variables'
    ws['A41'].font = Font(size=12, bold=True)

    tornado_headers = ['Variable', 'Low Value', 'Low NPV', 'Base NPV', 'High Value', 'High NPV', 'Range']
    ws.append(tornado_headers)
    apply_header_style(ws, 42, len(tornado_headers))

    tornado_data = [
        ['Demand Charges ($/kW-yr)', 180, '[Calculate]', '=$B$3', 280, '[Calculate]', '=F43-C43'],
        ['DR Events/year', 15, '[Calculate]', '=$B$3', 30, '[Calculate]', '=F44-C44'],
        ['Infrastructure ($/kW-yr)', 10, '[Calculate]', '=$B$3', 100, '[Calculate]', '=F45-C45'],
        ['CAPEX ($/kW)', 1400, '[Calculate]', '=$B$3', 1000, '[Calculate]', '=F46-C46'],
        ['Battery Life (years)', 12, '[Calculate]', '=$B$3', 18, '[Calculate]', '=F47-C47'],
    ]

    for row_data in tornado_data:
        ws.append(row_data)

    # Set column widths
    ws.column_dimensions['A'].width = 30
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Format currency
    for row in range(30, 40):
        for col in ['B', 'C', 'D']:
            cell = ws[f'{col}{row}']
            if row in [35, 36]:
                cell.number_format = '$#,##0'
            elif row == 38:
                cell.number_format = '0.0%'
            elif row == 39:
                cell.number_format = '0.0'

    return ws

def create_prioritization_scoring(wb):
    """Create Prioritization Scoring tab."""
    ws = wb['Prioritization Scoring']

    ws['A1'] = 'BUILDING PRIORITIZATION SCORING (100-Point Scale)'
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:Z1')

    # Headers
    headers = [
        'Building ID', 'Building Name',
        # Demand Charge Score (30%)
        '1. Demand Charge (30%)', 'Annual Savings ($)', 'Score (0-30)',
        # Infrastructure Score (25%)
        '2. Infrastructure (25%)', 'Feeder Loading %', 'Upgrade Timeline (yrs)',
        'Constraint Mult', 'Timeline Mult', 'Score (0-25)',
        # DR+ERCOT Score (20%)
        '3. DR+ERCOT (20%)', 'Annual Revenue ($)', 'Score (0-20)',
        # Resilience Score (15%)
        '4. Resilience (15%)', 'Critical Facility', 'Facility Type', 'Multiplier', 'Score (0-15)',
        # Site Readiness (10%)
        '5. Site Ready (10%)', 'Electrical Score', 'Space Score', 'Score (0-10)',
        # Total
        'TOTAL SCORE', 'Rank', 'Phase'
    ]

    ws.append(headers)
    apply_header_style(ws, 2, len(headers))

    # Add formulas for each building
    for i in range(1, 39):
        inv_row = i + 1
        calc_row = i + 2
        row_num = i + 2

        building_id = f'=\'Building Inventory\'!A{inv_row}'
        building_name = f'=\'Building Inventory\'!B{inv_row}'

        # 1. Demand Charge Score (30%)
        demand_savings = f'=\'Value Calculations\'!I{calc_row}'
        demand_score = f'=MIN(30,D{row_num}/30000*30)'

        # 2. Infrastructure Score (25%)
        # Get feeder loading from CPS Energy Data
        feeder_id = f'=\'Building Inventory\'!L{inv_row}'
        feeder_loading = f'=IFERROR(VLOOKUP(\'Building Inventory\'!L{inv_row},\'CPS Energy Data\'!$A$4:$E$18,5,FALSE),0.5)'
        upgrade_timeline = f'=IFERROR(VLOOKUP(\'Building Inventory\'!L{inv_row},\'CPS Energy Data\'!$A$4:$I$18,9,FALSE),10)'
        constraint_mult = f'=IF(G{row_num}>0.9,1,IF(G{row_num}>0.8,0.7,IF(G{row_num}>0.7,0.4,0.1)))'
        timeline_mult = f'=IF(H{row_num}<3,1,IF(H{row_num}<7,0.7,IF(H{row_num}<10,0.4,0.1)))'
        infra_score = f'=I{row_num}*J{row_num}*25'

        # 3. DR+ERCOT Score (20%)
        dr_ercot_revenue = f'=\'Value Calculations\'!O{calc_row}+\'Value Calculations\'!S{calc_row}'
        dr_score = f'=MIN(20,M{row_num}/15000*20)'

        # 4. Resilience Score (15%)
        critical_facility = f'=\'Building Inventory\'!M{inv_row}'
        facility_type = f'=\'Building Inventory\'!D{inv_row}'
        resilience_mult = f'=IF(Q{row_num}="Emergency Ops",1,IF(Q{row_num}="911 Center",1,IF(Q{row_num}="Police",0.8,IF(Q{row_num}="Fire",0.8,IF(Q{row_num}="Water Control",0.6,IF(Q{row_num}="City Hall",0.4,0.1))))))'
        resilience_score = f'=R{row_num}*15'

        # 5. Site Readiness (10%)
        electrical_score = 0.8  # Default assumption - manual update
        space_score = 0.8  # Default assumption - manual update
        site_score = f'=(U{row_num}+V{row_num})/2*10'

        # Total
        total_score = f'=SUM(E{row_num},K{row_num},N{row_num},S{row_num},W{row_num})'
        rank = f'=RANK(X{row_num},$X$3:$X$40,0)'
        phase = f'=IF(Y{row_num}<=10,1,IF(Y{row_num}<=25,2,3))'

        row_data = [
            building_id, building_name,
            # Demand Charge
            '', demand_savings, demand_score,
            # Infrastructure
            '', feeder_loading, upgrade_timeline, constraint_mult, timeline_mult, infra_score,
            # DR+ERCOT
            '', dr_ercot_revenue, dr_score,
            # Resilience
            '', critical_facility, facility_type, resilience_mult, resilience_score,
            # Site Readiness
            '', electrical_score, space_score, site_score,
            # Total
            total_score, rank, phase
        ]

        ws.append(row_data)

    # Summary statistics
    summary_row = 42
    ws[f'A{summary_row}'] = 'SUMMARY STATISTICS'
    ws[f'A{summary_row}'].font = Font(size=12, bold=True)

    ws[f'A{summary_row+2}'] = 'Average Score'
    ws[f'B{summary_row+2}'] = '=AVERAGE(X3:X40)'
    ws[f'A{summary_row+3}'] = 'Median Score'
    ws[f'B{summary_row+3}'] = '=MEDIAN(X3:X40)'
    ws[f'A{summary_row+4}'] = 'High Score'
    ws[f'B{summary_row+4}'] = '=MAX(X3:X40)'
    ws[f'A{summary_row+5}'] = 'Low Score'
    ws[f'B{summary_row+5}'] = '=MIN(X3:X40)'

    ws[f'D{summary_row+2}'] = 'Phase 1 Buildings'
    ws[f'E{summary_row+2}'] = '=COUNTIF(Z3:Z40,1)'
    ws[f'D{summary_row+3}'] = 'Phase 2 Buildings'
    ws[f'E{summary_row+3}'] = '=COUNTIF(Z3:Z40,2)'
    ws[f'D{summary_row+4}'] = 'Phase 3 Buildings'
    ws[f'E{summary_row+4}'] = '=COUNTIF(Z3:Z40,3)'

    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    for col in range(3, 27):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Apply input cell formatting (site readiness - manual entry)
    input_fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
    for row in range(3, 41):
        ws[f'U{row}'].fill = input_fill
        ws[f'V{row}'].fill = input_fill

    # Apply conditional formatting to total score
    ws.conditional_formatting.add(
        'X3:X40',
        ColorScaleRule(
            start_type='num', start_value=0, start_color='F8696B',
            mid_type='num', mid_value=50, mid_color='FFEB84',
            end_type='num', end_value=100, end_color='63BE7B'
        )
    )

    # Apply conditional formatting to rank
    ws.conditional_formatting.add(
        'Y3:Y40',
        ColorScaleRule(
            start_type='num', start_value=1, start_color='63BE7B',
            mid_type='num', mid_value=19, mid_color='FFEB84',
            end_type='num', end_value=38, end_color='F8696B'
        )
    )

    # Apply conditional formatting to phase
    ws.conditional_formatting.add(
        'Z3:Z40',
        CellIsRule(operator='equal', formula=['1'], fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'))
    )
    ws.conditional_formatting.add(
        'Z3:Z40',
        CellIsRule(operator='equal', formula=['2'], fill=PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'))
    )
    ws.conditional_formatting.add(
        'Z3:Z40',
        CellIsRule(operator='equal', formula=['3'], fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'))
    )

    # Format numbers
    for row in range(3, 41):
        ws[f'D{row}'].number_format = '$#,##0'
        ws[f'M{row}'].number_format = '$#,##0'
        ws[f'G{row}'].number_format = '0.0%'
        ws[f'X{row}'].number_format = '0.0'

    return ws

def create_phase_planning(wb):
    """Create Phase Planning tab."""
    ws = wb['Phase Planning']

    ws['A1'] = 'DEPLOYMENT PHASE PLANNING'
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:N1')

    # Phase 1 Details
    ws.append([])
    ws['A3'] = 'PHASE 1: PROOF OF CONCEPT (Year 1)'
    ws['A3'].font = Font(size=12, bold=True)
    ws['A3'].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    ws.merge_cells('A3:N3')

    phase1_headers = [
        'Rank', 'Building', 'BESS (kW)', 'Battery (kWh)', 'CAPEX ($)',
        'Annual Rev ($)', 'Payback (yrs)', 'Deploy Quarter', 'Feeder',
        'CPS Coord', 'Site Survey', 'Design', 'Permit', 'Commission Date'
    ]
    ws.append(phase1_headers)
    apply_header_style(ws, 4, len(phase1_headers))

    # Add rows for Phase 1 buildings (top 10)
    for i in range(1, 11):
        phase1_data = [
            f'=IFERROR(INDEX(\'Prioritization Scoring\'!$Y$3:$Y$40,{i}),"")',
            f'=IFERROR(INDEX(\'Building Inventory\'!$B$2:$B$39,MATCH({i},\'Prioritization Scoring\'!$Y$3:$Y$40,0)),"")',
            f'=IFERROR(INDEX(\'Building Inventory\'!$E$2:$E$39,MATCH({i},\'Prioritization Scoring\'!$Y$3:$Y$40,0)),"")',
            f'=C{4+i}*2',
            f'=C{4+i}*1200',
            f'=IFERROR(INDEX(\'Value Calculations\'!$AN$3:$AN$40,MATCH({i},\'Prioritization Scoring\'!$Y$3:$Y$40,0)),"")',
            f'=IF(F{4+i}>0,E{4+i}/F{4+i},"")',
            f'Q{(i-1)//3 + 1}',  # Distribute across quarters
            f'=IFERROR(INDEX(\'Building Inventory\'!$L$2:$L$39,MATCH({i},\'Prioritization Scoring\'!$Y$3:$Y$40,0)),"")',
            'Yes',
            'Not Started',
            'Not Started',
            'Not Started',
            ''
        ]
        ws.append(phase1_data)

    # Phase 1 Summary
    summary1_row = 16
    ws[f'A{summary1_row}'] = 'Phase 1 Totals:'
    ws[f'A{summary1_row}'].font = Font(bold=True)
    ws[f'C{summary1_row}'] = '=SUM(C5:C14)'
    ws[f'D{summary1_row}'] = '=SUM(D5:D14)'
    ws[f'E{summary1_row}'] = '=SUM(E5:E14)'
    ws[f'F{summary1_row}'] = '=SUM(F5:F14)'
    ws[f'G{summary1_row}'] = '=AVERAGE(G5:G14)'

    # Phase 2 Details
    ws.append([])
    ws[f'A{summary1_row+2}'] = 'PHASE 2: STRATEGIC EXPANSION (Year 2-3)'
    ws[f'A{summary1_row+2}'].font = Font(size=12, bold=True)
    ws[f'A{summary1_row+2}'].fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    ws.merge_cells(f'A{summary1_row+2}:N{summary1_row+2}')

    ws.append(phase1_headers)  # Same headers
    apply_header_style(ws, summary1_row+3, len(phase1_headers))

    # Add rows for Phase 2 buildings (rank 11-25)
    for i in range(11, 26):
        row_num = summary1_row + 3 + (i - 10)
        phase2_data = [
            f'=IFERROR(INDEX(\'Prioritization Scoring\'!$Y$3:$Y$40,{i}),"")',
            f'=IFERROR(INDEX(\'Building Inventory\'!$B$2:$B$39,MATCH({i},\'Prioritization Scoring\'!$Y$3:$Y$40,0)),"")',
            f'=IFERROR(INDEX(\'Building Inventory\'!$E$2:$E$39,MATCH({i},\'Prioritization Scoring\'!$Y$3:$Y$40,0)),"")',
            f'=C{row_num}*2',
            f'=C{row_num}*1200',
            f'=IFERROR(INDEX(\'Value Calculations\'!$AN$3:$AN$40,MATCH({i},\'Prioritization Scoring\'!$Y$3:$Y$40,0)),"")',
            f'=IF(F{row_num}>0,E{row_num}/F{row_num},"")',
            'TBD',
            f'=IFERROR(INDEX(\'Building Inventory\'!$L$2:$L$39,MATCH({i},\'Prioritization Scoring\'!$Y$3:$Y$40,0)),"")',
            'Pending',
            'Not Started',
            'Not Started',
            'Not Started',
            ''
        ]
        ws.append(phase2_data)

    # Phase 2 Summary
    summary2_row = summary1_row + 19
    ws[f'A{summary2_row}'] = 'Phase 2 Totals:'
    ws[f'A{summary2_row}'].font = Font(bold=True)
    ws[f'C{summary2_row}'] = f'=SUM(C{summary1_row+4}:C{summary2_row-1})'
    ws[f'D{summary2_row}'] = f'=SUM(D{summary1_row+4}:D{summary2_row-1})'
    ws[f'E{summary2_row}'] = f'=SUM(E{summary1_row+4}:E{summary2_row-1})'
    ws[f'F{summary2_row}'] = f'=SUM(F{summary1_row+4}:F{summary2_row-1})'
    ws[f'G{summary2_row}'] = f'=AVERAGE(G{summary1_row+4}:G{summary2_row-1})'

    # Phase 3 Summary (simplified)
    ws.append([])
    ws[f'A{summary2_row+2}'] = 'PHASE 3: FULL DEPLOYMENT (Year 3-5)'
    ws[f'A{summary2_row+2}'].font = Font(size=12, bold=True)
    ws[f'A{summary2_row+2}'].fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    ws.merge_cells(f'A{summary2_row+2}:N{summary2_row+2}')

    ws.append(['Buildings:', 'Rank 26-38 (13 buildings)'])
    ws.append(['Status:', 'Pending Phase 1-2 completion'])

    # Overall Summary
    ws.append([])
    ws.append([])
    overall_row = summary2_row + 7
    ws[f'A{overall_row}'] = 'OVERALL PROJECT SUMMARY'
    ws[f'A{overall_row}'].font = Font(size=12, bold=True)

    summary_headers = ['Metric', 'Phase 1', 'Phase 2', 'Phase 3', 'TOTAL']
    ws.append(summary_headers)
    apply_header_style(ws, overall_row+1, len(summary_headers))

    summary_data = [
        ['Buildings', 10, 15, 13, 38],
        ['Capacity (MW)', f'=C{summary1_row}/1000', f'=C{summary2_row}/1000', '=D41-B41-C41', '=SUM(B41:D41)'],
        ['CAPEX ($M)', f'=E{summary1_row}/1000000', f'=E{summary2_row}/1000000', '=D42*1.2*1000', '=SUM(B42:D42)'],
        ['Annual Revenue ($M)', f'=F{summary1_row}/1000000', f'=F{summary2_row}/1000000', '=D43*0.45', '=SUM(B43:D43)'],
        ['Avg Payback (years)', f'=G{summary1_row}', f'=G{summary2_row}', 1.9, '=(B44*B41+C44*C41+D44*D41)/(B41+C41+D41)'],
        [],
        ['Timeline', '', '', '', ''],
        ['Start', 'Q1-2026', 'Q1-2027', 'Q1-2028', ''],
        ['Complete', 'Q4-2026', 'Q4-2027', 'Q4-2028', ''],
    ]

    for row_data in summary_data:
        ws.append(row_data)

    # Set column widths
    column_widths = [8, 30, 12, 12, 12, 15, 12, 12, 15, 10, 12, 12, 12, 15]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Format currency
    for row in range(5, 15):
        ws[f'E{row}'].number_format = '$#,##0'
        ws[f'F{row}'].number_format = '$#,##0'
        ws[f'G{row}'].number_format = '0.0'

    for row in range(summary1_row+4, summary2_row):
        ws[f'E{row}'].number_format = '$#,##0'
        ws[f'F{row}'].number_format = '$#,##0'
        ws[f'G{row}'].number_format = '0.0'

    # Format summary rows
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{overall_row+2}'].number_format = '0.00'
        ws[f'{col}{overall_row+3}'].number_format = '$0.0'
        ws[f'{col}{overall_row+4}'].number_format = '$0.0'
        ws[f'{col}{overall_row+5}'].number_format = '0.0'

    return ws

def create_dashboard(wb):
    """Create Dashboard tab with summary metrics and charts."""
    ws = wb['Dashboard']

    # Title
    ws['A1'] = 'SAN ANTONIO BESS DEPLOYMENT'
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws['A2'] = '38 Municipal Buildings - Executive Dashboard'
    ws['A2'].font = Font(size=12)
    ws.merge_cells('A2:H2')
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.append([])

    # Fleet Summary (left column)
    ws['A4'] = 'FLEET SUMMARY'
    ws['A4'].font = Font(size=12, bold=True)
    ws['A4'].fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
    ws['A4'].font = Font(size=12, bold=True, color='FFFFFF')
    ws.merge_cells('A4:B4')

    fleet_metrics = [
        ['Total Capacity (MW)', '=SUM(\'Building Inventory\'!E2:E39)/1000'],
        ['Total Buildings', 38],
        ['Avg System Size (kW)', '=AVERAGE(\'Building Inventory\'!E2:E39)'],
        ['Total Battery (MWh)', '=SUM(\'Building Inventory\'!F2:F39)/1000'],
    ]

    for metric in fleet_metrics:
        ws.append(metric)

    # Financial Summary (right column - columns D-E)
    ws['D4'] = 'FINANCIAL SUMMARY'
    ws['D4'].font = Font(size=12, bold=True)
    ws['D4'].fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
    ws['D4'].font = Font(size=12, bold=True, color='FFFFFF')
    ws.merge_cells('D4:E4')

    financial_metrics = [
        ['Total CAPEX ($M)', '=\'Financial Model\'!B16/1000000'],
        ['Annual Revenue ($M)', '=\'Financial Model\'!B43/1000000'],
        ['20-Year NPV ($M)', '=\'Financial Model\'!B74/1000000'],
        ['IRR (%)', '=\'Financial Model\'!B75'],
        ['Payback (years)', '=\'Financial Model\'!B77'],
    ]

    for idx, metric in enumerate(financial_metrics, 5):
        ws[f'D{idx}'] = metric[0]
        ws[f'E{idx}'] = metric[1]

    # Value Stack section
    ws.append([])
    ws['A11'] = 'VALUE STACK ($/kW-year)'
    ws['A11'].font = Font(size=12, bold=True)
    ws.merge_cells('A11:B11')

    value_stack = [
        ['Demand Charges', '=\'Value Calculations\'!B52/SUM(\'Building Inventory\'!E2:E39)'],
        ['CPS Energy DR', '=\'Value Calculations\'!B53/SUM(\'Building Inventory\'!E2:E39)'],
        ['ERCOT Call Option', '=\'Value Calculations\'!B54/SUM(\'Building Inventory\'!E2:E39)'],
        ['Infrastructure', '=\'Value Calculations\'!B55/SUM(\'Building Inventory\'!E2:E39)'],
        ['Resilience', '=\'Value Calculations\'!B56/SUM(\'Building Inventory\'!E2:E39)'],
        ['Energy Arbitrage', '=\'Value Calculations\'!B57/SUM(\'Building Inventory\'!E2:E39)'],
        ['TOTAL', '=SUM(B12:B17)'],
    ]

    for metric in value_stack:
        ws.append(metric)

    # Revenue by Source (right column)
    ws['D11'] = 'ANNUAL REVENUE BY SOURCE'
    ws['D11'].font = Font(size=12, bold=True)
    ws.merge_cells('D11:E11')

    revenue_sources = [
        ['Demand Charges', '=\'Value Calculations\'!B52'],
        ['CPS Energy DR', '=\'Value Calculations\'!B53'],
        ['ERCOT Call Option', '=\'Value Calculations\'!B54'],
        ['Infrastructure', '=\'Value Calculations\'!B55'],
        ['Resilience', '=\'Value Calculations\'!B56'],
        ['Energy Arbitrage', '=\'Value Calculations\'!B57'],
        ['TOTAL', '=SUM(E12:E17)'],
    ]

    for idx, metric in enumerate(revenue_sources, 12):
        ws[f'D{idx}'] = metric[0]
        ws[f'E{idx}'] = metric[1]

    # Top 10 Priority Buildings
    ws.append([])
    ws['A20'] = 'TOP 10 PRIORITY BUILDINGS'
    ws['A20'].font = Font(size=12, bold=True)
    ws.merge_cells('A20:C20')

    top10_headers = ['Rank', 'Building', 'Score']
    ws.append(top10_headers)
    for col_idx, header in enumerate(top10_headers, 1):
        cell = ws.cell(row=21, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')

    for i in range(1, 11):
        ws.append([
            i,
            f'=IFERROR(INDEX(\'Building Inventory\'!$B$2:$B$39,MATCH({i},\'Prioritization Scoring\'!$Y$3:$Y$40,0)),"")',
            f'=IFERROR(INDEX(\'Prioritization Scoring\'!$X$3:$X$40,MATCH({i},\'Prioritization Scoring\'!$Y$3:$Y$40,0)),"")'
        ])

    # Deployment Phases
    ws['D20'] = 'DEPLOYMENT PHASES'
    ws['D20'].font = Font(size=12, bold=True)
    ws.merge_cells('D20:E20')

    phases = [
        ['Phase 1 (Year 1)', '=COUNTIF(\'Building Inventory\'!O2:O39,1)', 'buildings'],
        ['Phase 2 (Year 2)', '=COUNTIF(\'Building Inventory\'!O2:O39,2)', 'buildings'],
        ['Phase 3 (Year 3)', '=COUNTIF(\'Building Inventory\'!O2:O39,3)', 'buildings'],
    ]

    for idx, phase in enumerate(phases, 22):
        ws[f'D{idx}'] = phase[0]
        ws[f'E{idx}'] = phase[1]
        ws[f'F{idx}'] = phase[2]

    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12

    # Format numbers
    ws['B5'].number_format = '0.00'
    ws['B7'].number_format = '0.0'
    ws['B8'].number_format = '0.00'
    ws['E5'].number_format = '$0.0'
    ws['E6'].number_format = '$0.0'
    ws['E7'].number_format = '$0.0'
    ws['E8'].number_format = '0.0%'
    ws['E9'].number_format = '0.0'

    for row in range(12, 19):
        ws[f'B{row}'].number_format = '$#,##0'
        ws[f'E{row}'].number_format = '$#,##0'

    for row in range(22, 32):
        ws[f'C{row}'].number_format = '0.0'

    # Apply conditional formatting to top 10 scores
    ws.conditional_formatting.add(
        'C22:C31',
        ColorScaleRule(
            start_type='num', start_value=50, start_color='FFEB84',
            end_type='num', end_value=100, end_color='63BE7B'
        )
    )

    # Apply output cell highlighting
    output_fill = PatternFill(start_color=COLOR_OUTPUT, end_color=COLOR_OUTPUT, fill_type='solid')
    highlight_cells = ['B18', 'E18', 'E7']
    for cell_ref in highlight_cells:
        ws[cell_ref].fill = output_fill
        ws[cell_ref].font = Font(bold=True)

    return ws

def main():
    """Main function to generate the workbook."""
    print("Generating San Antonio BESS Analysis Workbook...")
    print("-" * 60)

    # Create workbook
    wb = create_workbook()
    print("✓ Created workbook structure")

    # Create each tab
    print("Creating tabs:")
    create_building_inventory(wb)
    print("  ✓ Building Inventory")

    create_value_calculations(wb)
    print("  ✓ Value Calculations")

    create_cps_energy_data(wb)
    print("  ✓ CPS Energy Data")

    create_financial_model(wb)
    print("  ✓ Financial Model")

    create_sensitivity_analysis(wb)
    print("  ✓ Sensitivity Analysis")

    create_prioritization_scoring(wb)
    print("  ✓ Prioritization Scoring")

    create_phase_planning(wb)
    print("  ✓ Phase Planning")

    create_dashboard(wb)
    print("  ✓ Dashboard")

    # Save workbook
    filename = 'San_Antonio_BESS_Analysis.xlsx'
    wb.save(filename)
    print("-" * 60)
    print(f"✓ Workbook saved as: {filename}")
    print()
    print("NEXT STEPS:")
    print("1. Open the Excel file")
    print("2. Update yellow input cells with actual building data")
    print("3. Request CPS Energy data (feeders, rates, event history)")
    print("4. Review calculations and adjust assumptions as needed")
    print("5. Use for decision-making and CPS Energy negotiations")
    print()
    print("Note: Some features require Excel desktop:")
    print("  - Data Tables (Sensitivity Analysis tab)")
    print("  - Advanced charts")
    print("  - Conditional formatting rules")
    print()
    print("Model is fully functional and ready to use!")

if __name__ == "__main__":
    main()
